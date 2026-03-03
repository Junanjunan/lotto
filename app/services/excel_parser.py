from __future__ import annotations

import io
from typing import Iterable

from openpyxl import load_workbook

from app.models import Draw


def _to_int(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        if not s:
            return None
        try:
            return int(float(s))
        except ValueError:
            return None
    return None


def _find_draw_fields(row: tuple) -> tuple[int, str, list[int], int] | None:
    values = [_to_int(v) for v in row]

    # Newer format: [row_no, draw_no, n1, n2, n3, n4, n5, n6, bonus, ...]
    if len(values) >= 9 and values[1] is not None and 1 <= values[1] <= 10000:
        draw_no = values[1]
        main = values[2:8]
        bonus = values[8]
        if len(main) == 6 and bonus is not None and _valid_lotto_numbers(main, bonus):
            return draw_no, "", sorted(main), bonus

    # Legacy/alternate format: [draw_no, n1, ..., n7, bonus, ...]
    if len(values) >= 8 and values[0] is not None and 1 <= values[0] <= 10000:
        draw_no = values[0]
        # if second column is date (YYYYMMDD), numbers start from index 2
        if values[1] is not None and values[1] > 20000000 and len(values) >= 9:
            main = values[2:8]
            bonus = values[8]
        else:
            # first column is draw_no and next 6 are main numbers
            main = values[1:7]
            if len(values) >= 8:
                bonus = values[7]
            else:
                bonus = None
        if len(main) == 6 and bonus is not None and _valid_lotto_numbers(main, bonus):
            return draw_no, "", sorted(main), bonus

    return None


def _valid_lotto_numbers(main: Iterable[int], bonus: int) -> bool:
    main_list = list(main)
    if len(main_list) != 6:
        return False
    if len(set(main_list)) != 6:
        return False
    if not all(1 <= n <= 45 for n in main_list):
        return False
    if not (1 <= bonus <= 45):
        return False
    return True


def parse_excel_draws(content: bytes) -> list[Draw]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    parsed: list[Draw] = []
    seen_draws: set[int] = set()

    for raw in rows:
        if all(v is None for v in raw):
            continue
        found = _find_draw_fields(raw)
        if not found:
            continue
        draw_no, draw_date, main_nums, bonus = found
        if draw_no in seen_draws:
            continue
        seen_draws.add(draw_no)
        parsed.append(Draw(draw_no=draw_no, draw_date=str(draw_date or ""), numbers=list(main_nums), bonus=bonus))

    return parsed
